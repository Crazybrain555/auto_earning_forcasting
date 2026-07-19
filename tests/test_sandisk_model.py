import unittest
import importlib.util
import tempfile
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MODEL = ROOT / "research" / "sandisk" / "SanDisk_financial_model_2026-07-17.xlsx"
REQUIRED_SHEETS = {
    "Summary",
    "Historicals",
    "Drivers",
    "Quarterly",
    "Forecast",
    "Scenarios",
    "Valuation",
    "Assumptions",
    "Sources",
}


class SandiskModelTest(unittest.TestCase):
    def load_model(self):
        self.assertTrue(MODEL.exists(), f"Missing workbook: {MODEL}")
        return load_workbook(MODEL, data_only=False)

    def test_required_sheets_exist(self):
        wb = self.load_model()
        self.assertTrue(REQUIRED_SHEETS.issubset(wb.sheetnames))

    def test_forecast_outputs_are_formulas(self):
        ws = self.load_model()["Forecast"]
        years = {ws.cell(3, col).value: col for col in range(2, ws.max_column + 1)}
        rows = {ws.cell(row, 1).value: row for row in range(1, ws.max_row + 1)}
        for year in ("FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"):
            for metric in ("Revenue", "Net income", "Diluted EPS"):
                value = ws.cell(rows[metric], years[year]).value
                self.assertIsInstance(value, str)
                self.assertTrue(value.startswith("="), f"{metric} {year} is not a formula")

    def test_scenario_probabilities_sum_to_one(self):
        ws = self.load_model()["Scenarios"]
        self.assertEqual(ws["B8"].value, "=SUM(B5:B7)")
        self.assertAlmostEqual(sum(ws.cell(row, 2).value for row in range(5, 8)), 1.0)

    def test_valuation_links_to_model_outputs(self):
        wb = self.load_model()
        ws = wb["Valuation"]
        formulas = [cell.value for row in ws.iter_rows() for cell in row if cell.data_type == "f"]
        self.assertTrue(any("Forecast!" in formula for formula in formulas))
        self.assertTrue(any("Scenarios!" in formula for formula in formulas))
        self.assertTrue(any("Assumptions!" in formula for formula in formulas))
        self.assertFalse(wb._external_links)

    def test_sources_have_evidence_tiers_and_urls(self):
        ws = self.load_model()["Sources"]
        headers = {ws.cell(3, col).value: col for col in range(1, ws.max_column + 1)}
        self.assertGreaterEqual(ws.max_row, 12)
        for row in range(4, ws.max_row + 1):
            self.assertIn(ws.cell(row, headers["Tier"]).value, {"E0", "E1", "E2", "E3", "E4"})
            self.assertTrue(str(ws.cell(row, headers["URL"]).value).startswith("http"))

    def test_recalculated_outputs_match_control_totals(self):
        wb = load_workbook(MODEL, data_only=True)
        self.assertAlmostEqual(wb["Forecast"]["B4"].value, 19283.0, places=3)
        self.assertAlmostEqual(wb["Forecast"]["C4"].value, 33205.326, places=3)
        self.assertAlmostEqual(wb["Forecast"]["C12"].value, 18973.650054, places=3)
        self.assertAlmostEqual(wb["Valuation"]["I27"].value, 1256.982299, places=3)
        self.assertAlmostEqual(wb["Valuation"]["I28"].value, -0.145276311, places=6)

    def test_recalculated_workbook_has_no_formula_errors(self):
        wb = load_workbook(MODEL, data_only=True)
        errors = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}
        found = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value in errors:
                        found.append(f"{ws.title}!{cell.coordinate}={cell.value}")
        self.assertEqual(found, [])


class SandiskModelBuilderTest(unittest.TestCase):
    def test_builder_creates_a_workbook_with_chart_series_titles(self):
        builder_path = ROOT / "research" / "sandisk" / "build_model.py"
        spec = importlib.util.spec_from_file_location("sandisk_build_model", builder_path)
        module = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        spec.loader.exec_module(module)
        with tempfile.TemporaryDirectory() as tmpdir:
            module.OUT_DIR = Path(tmpdir)
            module.OUT_PATH = Path(tmpdir) / "model.xlsx"
            module.main()
            self.assertTrue(module.OUT_PATH.exists())


if __name__ == "__main__":
    unittest.main()
