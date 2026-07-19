"""Build the project-local SanDisk FY2026E-FY2030E financial model.

The workbook is intentionally formula-driven.  Reported facts and explicit model
assumptions are blue hardcodes; same-sheet formulas are black; cross-sheet links
are green.  FY2026E is nine-month actual plus the midpoint of management's Q4
guidance and must not be described as a reported full-year result.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "research" / "sandisk"
OUT_PATH = OUT_DIR / "SanDisk_financial_model_2026-07-17.xlsx"
SOURCE_PATH = OUT_DIR / "source_manifest.csv"
INPUT_PATH = OUT_DIR / "model_inputs.json"


NAVY = "17365D"
MID_BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
PALE_GREEN = "E2F0D9"
PALE_YELLOW = "FFF2CC"
PALE_RED = "FCE4D6"
WHITE = "FFFFFF"
BLACK = "000000"
BLUE_INPUT = "0000FF"
GREEN_LINK = "008000"
GRAY = "666666"
GRID = Side(style="thin", color="D9E1F2")

FMT_MM = '#,##0.0;[Red](#,##0.0);-'
FMT_MM0 = '#,##0;[Red](#,##0);-'
FMT_PCT = '0.0%;[Red](0.0%);-'
FMT_EPS = '$0.00;[Red]($0.00);-'
FMT_PRICE = '$#,##0.00;[Red]($#,##0.00);-'
FMT_MULT = '0.0x'


SCENARIOS = {
    "Bear": {
        "probability": 0.30,
        "bit_growth": [0.12, 0.16, 0.15, 0.14],
        "asp_index": [75.0, 45.0, 35.0, 32.0],
        "cost_index": [98.0, 89.0, 82.0, 76.0],
        "gross_margin": [0.55, 0.30, 0.20, 0.28],
        "opex": [2300.0, 2200.0, 2100.0, 2200.0],
        "other_income": [100.0, 100.0, 100.0, 100.0],
        "tax_rate": [0.15, 0.15, 0.15, 0.16],
        "shares": [158.0, 158.0, 157.0, 156.0],
        "dc_share": [0.30, 0.33, 0.36, 0.39],
        "consumer_share": [0.17, 0.16, 0.15, 0.14],
        "quarterly_revenue": [6500.0, 5700.0, 5000.0, None],
        "quarterly_gm": [0.68, 0.58, 0.48, 0.42],
    },
    "Base": {
        "probability": 0.50,
        "bit_growth": [0.23, 0.23, 0.20, 0.18],
        "asp_index": [105.0, 80.0, 60.0, 50.0],
        "cost_index": [90.0, 79.0, 70.0, 62.0],
        "gross_margin": [0.74, 0.58, 0.46, 0.42],
        "opex": [2450.0, 2650.0, 2850.0, 3050.0],
        "other_income": [200.0, 300.0, 400.0, 500.0],
        "tax_rate": [0.15, 0.16, 0.17, 0.18],
        "shares": [154.0, 152.0, 151.0, 150.0],
        "dc_share": [0.35, 0.44, 0.49, 0.53],
        "consumer_share": [0.15, 0.13, 0.12, 0.11],
        "quarterly_revenue": [8600.0, 8800.0, 8300.0, None],
        "quarterly_gm": [0.80, 0.79, 0.72, 0.65],
    },
    "Bull": {
        "probability": 0.20,
        "bit_growth": [0.33, 0.30, 0.26, 0.23],
        "asp_index": [125.0, 118.0, 105.0, 90.0],
        "cost_index": [86.0, 73.0, 64.0, 55.0],
        "gross_margin": [0.80, 0.76, 0.68, 0.60],
        "opex": [2600.0, 3000.0, 3400.0, 3800.0],
        "other_income": [200.0, 350.0, 500.0, 700.0],
        "tax_rate": [0.15, 0.16, 0.17, 0.18],
        "shares": [153.0, 148.0, 144.0, 140.0],
        "dc_share": [0.42, 0.55, 0.61, 0.65],
        "consumer_share": [0.12, 0.10, 0.09, 0.08],
        "quarterly_revenue": [9500.0, 10200.0, 11000.0, None],
        "quarterly_gm": [0.82, 0.82, 0.81, 0.80],
    },
}


INLINE_SOURCES = [
    ("S1", "SEC", "SanDisk FY2025 Form 10-K", "2025-08-21", "E0", "FY2023-FY2025 audited financials and separation perimeter", "https://www.sec.gov/Archives/edgar/data/2023554/000202355425000034/sndk-20250627.htm", "Pre-separation periods are carve-out financial statements."),
    ("S2", "SanDisk IR", "FY2025 fourth-quarter and full-year results", "2025-08-14", "E1", "FY2025 non-GAAP bridge", "https://investor.sandisk.com/news-releases/news-release-details/sandisk-reports-fiscal-fourth-quarter-2025-financial-results", "Use the later 10-K for audited GAAP totals."),
    ("S3", "SanDisk IR", "FY2026 first-quarter results", "2025-11-06", "E1", "FY2026 Q1 actuals and BiCS8 stage", "https://investor.sandisk.com/news-releases/news-release-details/sandisk-reports-fiscal-first-quarter-2026-financial-results", "Company release; guidance is not actual."),
    ("S4", "SanDisk IR", "FY2026 second-quarter results", "2026-01-29", "E1", "FY2026 Q2 actuals and demand commentary", "https://investor.sandisk.com/news-releases/news-release-details/sandisk-reports-fiscal-second-quarter-2026-financial-results", "Company release; forward statements are conditional."),
    ("S5", "SanDisk IR", "FY2026 third-quarter results", "2026-04-30", "E1", "FY2026 Q3 actuals and Q4 guidance", "https://investor.sandisk.com/news-releases/news-release-details/sandisk-reports-fiscal-third-quarter-2026-financial-results", "Q4 is guidance only as of cutoff."),
    ("S6", "SEC", "SanDisk FY2026 Q3 Form 10-Q", "2026-05-01", "E0", "Latest statutory financials, RPO, cash and debt", "https://www.sec.gov/Archives/edgar/data/2023554/000162828026029401/sndk-20260403.htm", "Balance sheet predates the Nanya investment and possible buybacks."),
    ("S7", "SanDisk IR", "FY2026 Q3 earnings presentation", "2026-04-30", "E1", "Quarterly end-market and ASP/bit metrics", "https://investor.sandisk.com/static-files/8ea78860-f8e5-4f1c-ada3-c554437d6281", "Management presentation; non-GAAP definitions apply."),
    ("S8", "SanDisk IR", "BiCS10 1Tb TLC sampling announcement", "2026-07-02", "E1", "BiCS10 technical stage gate", "https://investor.sandisk.com/news-releases/news-release-details/sandisk-announces-sampling-bics10-1tb-tlc-3d-nand-flash-memory", "Sampling is not evidence of material revenue."),
    ("S9", "SanDisk IR", "FY2026 reporting date announcement", "2026-07-09", "E1", "Confirms Q4 results had not been reported by cutoff", "https://investor.sandisk.com/news-releases/news-release-details/sandisk-report-fiscal-fourth-quarter-and-fiscal-year-2026", "FY2026 results scheduled for 2026-08-05."),
    ("S10", "Kioxia", "Investor Day 2026 script", "2026-06-02", "E1", "NAND bit demand, capex and cost roadmap", "https://www.kioxia-holdings.com/content/dam/kioxia-hd/en-jp/ir/library/event/asset/Kioxia_Investor_Day_2026_en_script.pdf", "Peer plan; not SanDisk guidance."),
    ("S11", "Kioxia", "Investor Day 2026 Q&A", "2026-06-02", "E1", "Supply timing and product ramp", "https://www.kioxia-holdings.com/content/dam/kioxia-hd/en-jp/ir/library/event/asset/Investor-Day-2026-Eng-QA.pdf", "Peer statement; commercialization timing remains uncertain."),
    ("S12", "Kioxia", "Analyst Day 2025 presentation", "2025-09-30", "E1", "Flash Ventures capacity structure and cost roadmap", "https://www.kioxia-holdings.com/content/dam/kioxia-hd/en-jp/ir/event/asset/Kioxia-Analyst-Day-202509-en.pdf", "Kioxia framing; apply to SanDisk only through disclosed JV economics."),
    ("S13", "Kioxia", "K2 BiCS10 production announcement", "2026-07-03", "E1", "BiCS10 initial production evidence", "https://www.kioxia.com/en-jp/about/news/2026/20260703-2.html", "Initial production does not establish high-volume output or revenue."),
    ("S14", "Micron", "FY2026 Q3 prepared remarks", "2026-06-24", "E1", "Industry NAND bit-demand and tightness outlook", "https://investors.micron.com/static-files/631b1a32-5537-46ae-8f40-82e42fc79dfe", "Peer outlook, not independent industry fact."),
    ("S15", "TrendForce", "NAND supplier survey", "2026-05-25", "E3", "NAND supply and pricing context", "https://www.trendforce.com/presscenter/news/20260525-13058.html", "Third-party estimate; methodology is not fully transparent."),
    ("S16", "TrendForce", "Enterprise SSD market survey", "2026-06-11", "E3", "Enterprise SSD demand and pricing", "https://www.trendforce.com/presscenter/news/20260611-13092.html", "Third-party estimate; may be revised."),
    ("S17", "TrendForce", "2026 NAND capex outlook", "2025-11-13", "E3", "Industry capacity discipline", "https://www.trendforce.com/presscenter/news/20251113-12780.html", "Industry view does not eliminate K2-specific supply additions."),
    ("S18", "Nasdaq", "SNDK real-time quote API", "2026-07-17", "E0", "Intraday price at 12:23 ET", "https://api.nasdaq.com/api/quote/SNDK/info?assetclass=stocks", "Intraday quote, not a closing price."),
    ("S19", "Nasdaq", "SNDK historical quote API", "2026-07-17", "E0", "Most recent completed close on 2026-07-16", "https://api.nasdaq.com/api/quote/SNDK/historical?assetclass=stocks&fromdate=2026-07-16&todate=2026-07-17&limit=10", "Historical API field conventions should be checked before automation."),
    ("S20", "Nasdaq", "SNDK analyst earnings forecast API", "2026-07-17", "E3", "Small-sample consensus cross-check", "https://api.nasdaq.com/api/analyst/SNDK/earnings-forecast", "FY2028 contained only one estimate at retrieval."),
]


def add_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color=GRAY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 28


def section(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Arial", bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=MID_BLUE)


def header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for idx, value in enumerate(headers, start_col):
        cell = ws.cell(row, idx, value)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=GRID)


def put(ws, row: int, col: int, value, number_format: str | None = None, comment: str | None = None):
    cell = ws.cell(row, col, value)
    cell.font = Font(name="Arial", size=10, color=BLACK)
    if isinstance(value, str) and value.startswith("="):
        color = GREEN_LINK if "!" in value else BLACK
        cell.font = Font(name="Arial", size=10, color=color)
    elif isinstance(value, (int, float)):
        cell.font = Font(name="Arial", size=10, color=BLUE_INPUT)
    if number_format:
        cell.number_format = number_format
    if comment:
        cell.comment = Comment(comment, "Codex")
    return cell


def finish_sheet(ws, widths: dict[int, float], freeze: str = "B4") -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.font.name != "Arial":
                cell.font = Font(
                    name="Arial",
                    size=cell.font.sz or 10,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                )
            if cell.value is not None:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical or "center",
                    wrap_text=cell.alignment.wrap_text,
                )


def build_assumptions(wb: Workbook) -> None:
    ws = wb.create_sheet("Assumptions")
    add_title(ws, "SanDisk assumptions register", "Blue cells are explicit inputs; model formulas reference this sheet.", 6)
    header_row(ws, 3, ["Assumption", "Value", "Unit", "Type", "Source / rationale", "Trigger or limitation"])
    rows = [
        ("Current price", 1470.63, "$/share", "Observed", "S18; Nasdaq 2026-07-17 12:23 ET", "Intraday, not close"),
        ("Basic shares outstanding", 148.089758, "mm", "Reported", "S6; as of 2026-04-24", "May change after buybacks"),
        ("Cash", 3735.0, "$mm", "Reported", "S6; 2026-04-03", "Predates Nanya investment"),
        ("Debt", 0.0, "$mm", "Reported", "S6; term loan repaid", "Flash Ventures commitments remain"),
        ("Nanya cash payment", 972.0, "$mm", "Reported", "S6 subsequent event", "Investment can also be treated as non-operating asset"),
        ("Last completed close", 1411.08, "$/share", "Observed", "S19; 2026-07-16", "Included as reference only"),
        ("Q4 FY26 revenue midpoint", 8000.0, "$mm", "Guidance", "S5; range $7.75-$8.25bn", "Not reported actual"),
        ("Q4 FY26 GAAP gross margin midpoint", 0.799, "%", "Guidance", "S5; range 78.9%-80.9%", "Not reported actual"),
        ("Q4 FY26 GAAP opex midpoint", 540.5, "$mm", "Guidance", "S5; range $523-$558mm", "Not reported actual"),
        ("Q4 FY26 other income midpoint", 22.0, "$mm", "Guidance", "S5; range $12-$32mm", "Not reported actual"),
        ("FY26E full-year tax rate", 0.143, "%", "Model", "Calibrated from 9M actual; Q4 GAAP tax not guided", "Sensitivity required"),
        ("FY26E diluted shares", 156.0, "mm", "Model", "9M 154mm; Q4 guide about 158mm", "Buyback execution unknown"),
        ("FY26 average ASP index", 75.0, "index", "Model", "Q3 FY26 = 100 anchor", "No absolute ASP/GB disclosed"),
        ("Q4 FY26 Datacenter revenue share", 0.30, "%", "Model", "Q3 mix plus enterprise demand", "No company Q4 segment guide"),
        ("Q4 FY26 Edge revenue share", 0.58, "%", "Model", "Residual demand mix", "No company Q4 segment guide"),
        ("Q4 FY26 Consumer revenue share", 0.12, "%", "Model", "Q3 mix and price elasticity", "No company Q4 segment guide"),
        ("FY26E D&A proxy", 150.0, "$mm", "Model", "9M cash-flow disclosure plus Q4 estimate", "JV structure limits peer comparability"),
        ("Selected scenario", "Base", "text", "Model", "Central operating path", "Change only with explicit model revision"),
        ("Transition valuation weight", 0.40, "%", "Model", "Balances FY27 peak/transition earnings", "Not a market fact"),
        ("Normalized valuation weight", 0.60, "%", "Model", "Balances FY29-FY30 earnings", "Not a market fact"),
        ("FY26E non-GAAP EPS", 62.82, "$/share", "Derived", "9M $31.32 + Q4 guide midpoint $31.50", "Mixed actual and guidance"),
        ("Q4 FY26 diluted shares guide", 158.0, "mm", "Guidance", "S5; approximately 158mm", "Not reported actual"),
    ]
    for r, values in enumerate(rows, 4):
        for c, value in enumerate(values, 1):
            fmt = None
            if c == 2 and isinstance(value, (int, float)):
                if values[2] == "%":
                    fmt = FMT_PCT
                elif values[2] == "$/share":
                    fmt = FMT_PRICE
                else:
                    fmt = FMT_MM
            put(ws, r, c, value, fmt)

    section(ws, 28, "Scenario valuation assumptions", 6)
    header_row(ws, 29, ["Metric", "Bear", "Base", "Bull", "Unit", "Rationale"])
    valuation_rows = [
        ("Transition P/E", 8.0, 10.0, 14.0, "x", "Peak-cycle earnings receive a lower multiple"),
        ("Normalized P/E", 10.0, 16.0, 18.0, "x", "Long-term contract validation and balance sheet support the range"),
        ("Excess cash / share", 10.0, 50.0, 100.0, "$/share", "Scenario treatment of cash, buyback and strategic investment"),
    ]
    for r, values in enumerate(valuation_rows, 30):
        for c, value in enumerate(values, 1):
            fmt = FMT_MULT if r < 32 and 2 <= c <= 4 else (FMT_PRICE if r == 32 and 2 <= c <= 4 else None)
            put(ws, r, c, value, fmt)
    finish_sheet(ws, {1: 38, 2: 16, 3: 14, 4: 14, 5: 50, 6: 48}, "B4")


def build_historicals(wb: Workbook, inputs: dict) -> None:
    ws = wb.create_sheet("Historicals")
    add_title(ws, "SanDisk historical financials", "$mm except per-share data; FY2026E is 9M actual plus Q4 guidance midpoint.", 6)
    header_row(ws, 3, ["Metric", "FY2023A", "FY2024A", "FY2025A", "FY2026 9M A", "FY2026E"])
    periods = [
        inputs["historicals"]["FY2023"],
        inputs["historicals"]["FY2024"],
        inputs["historicals"]["FY2025"],
        inputs["quarterly"]["FY2026_9M"],
    ]
    data = {
        "Revenue": [*[p["revenue"] for p in periods], "=E4+Assumptions!B10"],
        "Gross profit": [*[p["gross_profit"] for p in periods], "=E5+Assumptions!B10*Assumptions!B11"],
        "Gross margin": ["=B5/B4", "=C5/C4", "=D5/D4", "=E5/E4", "=F5/F4"],
        "Operating expenses": [*[p["total_operating_expenses"] for p in periods], "=E7+Assumptions!B12"],
        "Operating income": [*[p["operating_income"] for p in periods], "=F5-F7"],
        "Interest and other, net": [*[p["other_income_expense"] for p in periods], "=E9+Assumptions!B13"],
        "Pretax income": [*[p["income_before_tax"] for p in periods], "=F8+F9"],
        "Income taxes": [*[p["income_tax_expense"] for p in periods], "=MAX(0,F10*Assumptions!B14)"],
        "Net income": [*[p["net_income"] for p in periods], "=F10-F11"],
        "Diluted shares": [*[p["diluted_shares"] for p in periods], "=Assumptions!B15"],
        "Diluted EPS": [*[p["diluted_eps"] for p in periods], "=F12/F13"],
        "Operating cash flow": [*[p["cash_from_operations"] for p in periods], None],
        "PP&E purchases": [*[p["gross_capex"] for p in periods], None],
        "": [None, None, None, None, None],
        "Datacenter / Cloud revenue": [*[p["end_markets"]["Datacenter"] for p in periods], "=E18+Assumptions!B10*Assumptions!B17"],
        "Edge / Client revenue": [*[p["end_markets"]["Edge"] for p in periods], "=E19+Assumptions!B10*Assumptions!B18"],
        "Consumer revenue": [*[p["end_markets"]["Consumer"] for p in periods], "=E20+Assumptions!B10*Assumptions!B19"],
    }
    for r, (metric, values) in enumerate(data.items(), 4):
        put(ws, r, 1, metric)
        for idx, value in enumerate(values, 2):
            fmt = FMT_PCT if metric == "Gross margin" else (FMT_EPS if metric == "Diluted EPS" else FMT_MM)
            comment = None
            if metric == "Revenue" and idx == 6:
                comment = "FY2026E combines reported 9M actual with management's Q4 revenue guidance midpoint; it is not a reported actual."
            put(ws, r, idx, value, fmt, comment)
    ws["A22"] = "Perimeter note"
    ws["B22"] = "FY2023-FY2025 include carve-out allocations from Western Digital; pre-separation operating expenses are not fully comparable with standalone SanDisk."
    ws.merge_cells("B22:F22")
    ws["A22"].font = Font(name="Arial", bold=True, color=NAVY)
    ws["B22"].font = Font(name="Arial", italic=True, color=GRAY)
    ws["B22"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[22].height = 38
    finish_sheet(ws, {1: 34, 2: 15, 3: 15, 4: 15, 5: 18, 6: 16}, "B4")


def build_drivers(wb: Workbook) -> dict[str, int]:
    ws = wb.create_sheet("Drivers")
    add_title(ws, "SanDisk scenario drivers", "Q3 FY2026 ASP/GB and cost/GB are indexed to 100. All FY2027-FY2030 inputs are explicit scenarios, not guidance.", 5)
    header_row(ws, 3, ["Driver", "FY2027E", "FY2028E", "FY2029E", "FY2030E"])
    row_map: dict[str, int] = {}
    start_rows = {"Bear": 4, "Base": 17, "Bull": 30}
    labels = [
        ("bit_growth", "Bit shipment growth", FMT_PCT),
        ("asp_index", "ASP / GB index", '0.0'),
        ("cost_index", "Cost / GB index", '0.0'),
        ("gross_margin", "Gross margin", FMT_PCT),
        ("opex", "Operating expenses", FMT_MM),
        ("other_income", "Interest and other income", FMT_MM),
        ("tax_rate", "Tax rate", FMT_PCT),
        ("shares", "Diluted shares", FMT_MM),
        ("dc_share", "Datacenter revenue share", FMT_PCT),
        ("edge_share", "Edge revenue share", FMT_PCT),
        ("consumer_share", "Consumer revenue share", FMT_PCT),
    ]
    for scenario, start in start_rows.items():
        section(ws, start, scenario, 5)
        for offset, (key, label, fmt) in enumerate(labels, 1):
            row = start + offset
            row_map[f"{scenario}_{key}"] = row
            put(ws, row, 1, label)
            if key == "edge_share":
                values = [1 - d - c for d, c in zip(SCENARIOS[scenario]["dc_share"], SCENARIOS[scenario]["consumer_share"])]
            else:
                values = SCENARIOS[scenario][key]
            for col, value in enumerate(values, 2):
                put(ws, row, col, value, fmt)
        for col in range(2, 6):
            ws.cell(start, col).fill = PatternFill("solid", fgColor=MID_BLUE)
    ws["A43"] = "Driver logic"
    ws["B43"] = "Revenue = prior-year revenue x (1 + bit growth) x (current ASP index / prior ASP index). Gross margin is separately calibrated to ASP, cost/GB, mix and utilization because absolute company cost/GB is not disclosed."
    ws.merge_cells("B43:E43")
    ws["B43"].alignment = Alignment(wrap_text=True)
    ws["B43"].font = Font(name="Arial", italic=True, color=GRAY)
    ws.row_dimensions[43].height = 52
    finish_sheet(ws, {1: 34, 2: 16, 3: 16, 4: 16, 5: 16}, "B4")
    return row_map


def scenario_revenue_formula(scenario: str, year_index: int, scenario_row: int, row_map: dict[str, int]) -> str:
    driver_col = get_column_letter(2 + year_index)
    bit_row = row_map[f"{scenario}_bit_growth"]
    asp_row = row_map[f"{scenario}_asp_index"]
    if year_index == 0:
        return f"=Forecast!$B$4*(1+Drivers!{driver_col}{bit_row})*(Drivers!{driver_col}{asp_row}/Assumptions!$B$16)"
    prior_col = get_column_letter(1 + year_index)
    prior_asp_col = get_column_letter(1 + year_index)
    return f"={prior_col}{scenario_row}*(1+Drivers!{driver_col}{bit_row})*(Drivers!{driver_col}{asp_row}/Drivers!{prior_asp_col}{asp_row})"


def build_scenarios(wb: Workbook, row_map: dict[str, int]) -> dict[str, dict[str, int]]:
    ws = wb.create_sheet("Scenarios")
    add_title(ws, "SanDisk bear / base / bull scenarios", "$mm except per-share data. Probabilities are explicit judgment weights, not statistical frequencies.", 6)
    header_row(ws, 4, ["Scenario", "Probability"])
    for row, scenario in enumerate(["Bear", "Base", "Bull"], 5):
        put(ws, row, 1, scenario)
        put(ws, row, 2, SCENARIOS[scenario]["probability"], FMT_PCT)
    put(ws, 8, 1, "Total")
    put(ws, 8, 2, "=SUM(B5:B7)", FMT_PCT)
    ws["A8"].font = Font(name="Arial", bold=True)
    ws["B8"].font = Font(name="Arial", bold=True, color=BLACK)

    tables = {
        "Revenue": (10, 11, FMT_MM),
        "Net income": (16, 17, FMT_MM),
        "Diluted EPS": (22, 23, FMT_EPS),
    }
    positions: dict[str, dict[str, int]] = {metric: {} for metric in tables}
    years = ["FY2027E", "FY2028E", "FY2029E", "FY2030E"]
    for metric, (title_row, header, fmt) in tables.items():
        section(ws, title_row, metric, 5)
        header_row(ws, header, ["Scenario", *years])
        for i, scenario in enumerate(["Bear", "Base", "Bull"], header + 1):
            positions[metric][scenario] = i
            put(ws, i, 1, scenario)
            for year_idx in range(4):
                col = 2 + year_idx
                if metric == "Revenue":
                    formula = scenario_revenue_formula(scenario, year_idx, i, row_map)
                elif metric == "Net income":
                    revenue_row = positions["Revenue"][scenario]
                    rev_ref = f"{get_column_letter(col)}{revenue_row}"
                    dcol = get_column_letter(col)
                    gm_r = row_map[f"{scenario}_gross_margin"]
                    op_r = row_map[f"{scenario}_opex"]
                    oi_r = row_map[f"{scenario}_other_income"]
                    tax_r = row_map[f"{scenario}_tax_rate"]
                    pretax = f"({rev_ref}*Drivers!{dcol}{gm_r}-Drivers!{dcol}{op_r}+Drivers!{dcol}{oi_r})"
                    formula = f"={pretax}-MAX(0,{pretax}*Drivers!{dcol}{tax_r})"
                else:
                    ni_row = positions["Net income"][scenario]
                    dcol = get_column_letter(col)
                    share_r = row_map[f"{scenario}_shares"]
                    formula = f"={dcol}{ni_row}/Drivers!{dcol}{share_r}"
                put(ws, i, col, formula, fmt)
    ws["A28"] = "Interpretation"
    ws["B28"] = "Bear is a supply-led price reset; Base assumes FY27 tightness followed by price normalization; Bull assumes durable AI/datacenter scarcity plus successful enterprise SSD execution."
    ws.merge_cells("B28:F28")
    ws["B28"].alignment = Alignment(wrap_text=True)
    ws["B28"].font = Font(name="Arial", italic=True, color=GRAY)
    ws.row_dimensions[28].height = 42
    finish_sheet(ws, {1: 20, 2: 16, 3: 16, 4: 16, 5: 16, 6: 12}, "B5")
    return positions


def build_forecast(wb: Workbook, row_map: dict[str, int]) -> None:
    ws = wb.create_sheet("Forecast")
    add_title(ws, "SanDisk base-case annual forecast", "$mm except per-share data. FY2026E is mixed actual/guidance; FY2027E-FY2030E are explicit model estimates.", 6)
    header_row(ws, 3, ["Metric", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"])
    metrics = [
        "Revenue",
        "Gross profit",
        "Gross margin",
        "Operating expenses",
        "Operating income",
        "Interest and other income",
        "Pretax income",
        "Income taxes",
        "Net income",
        "Diluted shares",
        "Diluted EPS",
        "EBITDA proxy",
        "Datacenter revenue",
        "Edge revenue",
        "Consumer revenue",
    ]
    row_for = {metric: row for row, metric in enumerate(metrics, 4)}
    for metric, row in row_for.items():
        put(ws, row, 1, metric)

    fy26_formulas = {
        "Revenue": "=Historicals!F4",
        "Gross profit": "=Historicals!F5",
        "Gross margin": "=B5/B4",
        "Operating expenses": "=Historicals!F7",
        "Operating income": "=B5-B7",
        "Interest and other income": "=Historicals!F9",
        "Pretax income": "=B8+B9",
        "Income taxes": "=Historicals!F11",
        "Net income": "=B10-B11",
        "Diluted shares": "=Historicals!F13",
        "Diluted EPS": "=B12/B13",
        "EBITDA proxy": "=B8+Assumptions!B20",
        "Datacenter revenue": "=Historicals!F18",
        "Edge revenue": "=Historicals!F19",
        "Consumer revenue": "=Historicals!F20",
    }
    for metric, formula in fy26_formulas.items():
        fmt = FMT_PCT if metric == "Gross margin" else (FMT_EPS if metric == "Diluted EPS" else FMT_MM)
        put(ws, row_for[metric], 2, formula, fmt)

    for year_idx in range(4):
        col = 3 + year_idx
        col_letter = get_column_letter(col)
        driver_col = get_column_letter(2 + year_idx)
        prior_col = get_column_letter(col - 1)
        bit_row = row_map["Base_bit_growth"]
        asp_row = row_map["Base_asp_index"]
        if year_idx == 0:
            revenue_formula = f"=B4*(1+Drivers!{driver_col}{bit_row})*(Drivers!{driver_col}{asp_row}/Assumptions!B16)"
        else:
            prior_driver_col = get_column_letter(1 + year_idx)
            revenue_formula = f"={prior_col}4*(1+Drivers!{driver_col}{bit_row})*(Drivers!{driver_col}{asp_row}/Drivers!{prior_driver_col}{asp_row})"
        formulas = {
            "Revenue": revenue_formula,
            "Gross profit": f"={col_letter}4*{col_letter}6",
            "Gross margin": f"=Drivers!{driver_col}{row_map['Base_gross_margin']}",
            "Operating expenses": f"=Drivers!{driver_col}{row_map['Base_opex']}",
            "Operating income": f"={col_letter}5-{col_letter}7",
            "Interest and other income": f"=Drivers!{driver_col}{row_map['Base_other_income']}",
            "Pretax income": f"={col_letter}8+{col_letter}9",
            "Income taxes": f"=MAX(0,{col_letter}10*Drivers!{driver_col}{row_map['Base_tax_rate']})",
            "Net income": f"={col_letter}10-{col_letter}11",
            "Diluted shares": f"=Drivers!{driver_col}{row_map['Base_shares']}",
            "Diluted EPS": f"={col_letter}12/{col_letter}13",
            "EBITDA proxy": f"={col_letter}8+Assumptions!$B$20",
            "Datacenter revenue": f"={col_letter}4*Drivers!{driver_col}{row_map['Base_dc_share']}",
            "Edge revenue": f"={col_letter}4*Drivers!{driver_col}{row_map['Base_edge_share']}",
            "Consumer revenue": f"={col_letter}4*Drivers!{driver_col}{row_map['Base_consumer_share']}",
        }
        for metric, formula in formulas.items():
            fmt = FMT_PCT if metric == "Gross margin" else (FMT_EPS if metric == "Diluted EPS" else FMT_MM)
            put(ws, row_for[metric], col, formula, fmt)
    ws["A21"] = "Caveat"
    ws["B21"] = "The gross-margin path is an explicit cycle/contract/mix calibration. Q3/Q4 FY2026 scarcity margins are not treated as a permanent cost structure."
    ws.merge_cells("B21:F21")
    ws["B21"].alignment = Alignment(wrap_text=True)
    ws["B21"].font = Font(name="Arial", italic=True, color=GRAY)
    ws.row_dimensions[21].height = 42
    finish_sheet(ws, {1: 34, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16}, "B4")


def build_quarterly(wb: Workbook, row_map: dict[str, int], positions: dict[str, dict[str, int]]) -> None:
    ws = wb.create_sheet("Quarterly")
    add_title(ws, "FY2027 quarterly scenario schedule", "$mm except per-share data. Quarterly shapes are model assumptions and reconcile to each FY2027 scenario.", 5)
    header_row(ws, 3, ["Metric", "Q1 FY27E", "Q2 FY27E", "Q3 FY27E", "Q4 FY27E"])
    starts = {"Bear": 4, "Base": 16, "Bull": 28}
    for scenario, start in starts.items():
        section(ws, start, scenario, 5)
        rows = {
            "Revenue": start + 1,
            "Gross margin": start + 2,
            "Gross profit": start + 3,
            "Operating expenses": start + 4,
            "Operating income": start + 5,
            "Other income": start + 6,
            "Pretax income": start + 7,
            "Income taxes": start + 8,
            "Net income": start + 9,
            "Diluted EPS": start + 10,
        }
        for label, row in rows.items():
            put(ws, row, 1, label)
        for q in range(4):
            col = 2 + q
            c = get_column_letter(col)
            if q < 3:
                revenue = SCENARIOS[scenario]["quarterly_revenue"][q]
            else:
                annual_rev_row = positions["Revenue"][scenario]
                revenue = f"=Scenarios!B{annual_rev_row}-SUM(B{rows['Revenue']}:D{rows['Revenue']})"
            put(ws, rows["Revenue"], col, revenue, FMT_MM)
            put(ws, rows["Gross margin"], col, SCENARIOS[scenario]["quarterly_gm"][q], FMT_PCT)
            put(ws, rows["Gross profit"], col, f"={c}{rows['Revenue']}*{c}{rows['Gross margin']}", FMT_MM)
            put(ws, rows["Operating expenses"], col, f"=Drivers!B{row_map[f'{scenario}_opex']}/4", FMT_MM)
            put(ws, rows["Operating income"], col, f"={c}{rows['Gross profit']}-{c}{rows['Operating expenses']}", FMT_MM)
            put(ws, rows["Other income"], col, f"=Drivers!B{row_map[f'{scenario}_other_income']}/4", FMT_MM)
            put(ws, rows["Pretax income"], col, f"={c}{rows['Operating income']}+{c}{rows['Other income']}", FMT_MM)
            put(ws, rows["Income taxes"], col, f"=MAX(0,{c}{rows['Pretax income']}*Drivers!B{row_map[f'{scenario}_tax_rate']})", FMT_MM)
            put(ws, rows["Net income"], col, f"={c}{rows['Pretax income']}-{c}{rows['Income taxes']}", FMT_MM)
            put(ws, rows["Diluted EPS"], col, f"={c}{rows['Net income']}/Drivers!B{row_map[f'{scenario}_shares']}", FMT_EPS)
    finish_sheet(ws, {1: 32, 2: 16, 3: 16, 4: 16, 5: 16}, "B4")


def build_valuation(wb: Workbook, positions: dict[str, dict[str, int]]) -> None:
    ws = wb.create_sheet("Valuation")
    add_title(ws, "SanDisk current and normalized valuation", "$mm except per-share data. Current price is 2026-07-17 12:23 ET intraday, not a close.", 9)
    header_row(ws, 3, ["Metric", "Value", "Unit", "Definition"])
    current = [
        ("Current price", "=Assumptions!B4", "$/share", "Nasdaq intraday"),
        ("Basic shares outstanding", "=Assumptions!B5", "mm", "SEC actual shares"),
        ("Market capitalization", "=B4*B5", "$mm", "Price x basic shares"),
        ("Cash", "=Assumptions!B6", "$mm", "Latest reported cash"),
        ("Debt", "=Assumptions!B7", "$mm", "Latest reported debt"),
        ("Enterprise value", "=B6+B8-B7", "$mm", "Market cap + debt - cash"),
        ("Pro forma EV after Nanya cash payment", "=B9+Assumptions!B8", "$mm", "Conservative cash-only treatment"),
        ("FY2026E revenue", "=Forecast!B4", "$mm", "9M actual + Q4 guide midpoint"),
        ("FY2026E GAAP net income", "=Forecast!B12", "$mm", "Model estimate"),
        ("FY2026E GAAP diluted EPS", "=Forecast!B14", "$/share", "Model estimate"),
        ("FY2026E non-GAAP diluted EPS", "=Assumptions!B24", "$/share", "9M actual + Q4 guide midpoint"),
        ("FY2026E EBITDA proxy", "=Forecast!B15", "$mm", "Operating income + D&A proxy"),
        ("Current P/E on GAAP FY2026E", "=B4/B13", "x", "Price / model EPS"),
        ("Current P/E on non-GAAP FY2026E", "=B4/B14", "x", "Price / mixed actual-guide EPS"),
        ("EV / sales", "=B9/B11", "x", "EV / FY2026E revenue"),
        ("EV / EBITDA proxy", "=B9/B15", "x", "JV accounting limits peer comparison"),
    ]
    for r, values in enumerate(current, 4):
        for c, value in enumerate(values, 1):
            fmt = None
            if c == 2:
                if r == 4:
                    fmt = FMT_PRICE
                elif r in (13, 14):
                    fmt = FMT_EPS
                elif r >= 16:
                    fmt = FMT_MULT
                else:
                    fmt = FMT_MM
            put(ws, r, c, value, fmt)

    section(ws, 22, "Scenario fair value", 9)
    header_row(ws, 23, ["Scenario", "Probability", "FY27 EPS", "Transition P/E", "FY29-30 avg EPS", "Normalized P/E", "Excess cash/share", "Fair value/share", "Probability-weighted"])
    scenario_cols = {"Bear": "B", "Base": "C", "Bull": "D"}
    probability_rows = {"Bear": 5, "Base": 6, "Bull": 7}
    for r, scenario in enumerate(["Bear", "Base", "Bull"], 24):
        eps_row = positions["Diluted EPS"][scenario]
        assumption_col = scenario_cols[scenario]
        values = [
            scenario,
            f"=Scenarios!B{probability_rows[scenario]}",
            f"=Scenarios!B{eps_row}",
            f"=Assumptions!{assumption_col}30",
            f"=AVERAGE(Scenarios!D{eps_row}:E{eps_row})",
            f"=Assumptions!{assumption_col}31",
            f"=Assumptions!{assumption_col}32",
            f"=C{r}*D{r}*Assumptions!$B$22+E{r}*F{r}*Assumptions!$B$23+G{r}",
            f"=B{r}*H{r}",
        ]
        for c, value in enumerate(values, 1):
            fmt = FMT_PCT if c == 2 else (FMT_MULT if c in (4, 6) else (FMT_PRICE if c >= 3 else None))
            put(ws, r, c, value, fmt)
    put(ws, 27, 1, "Probability-weighted fair value")
    put(ws, 27, 9, "=SUM(I24:I26)", FMT_PRICE)
    put(ws, 28, 1, "Upside / (downside) to intraday price")
    put(ws, 28, 9, "=I27/B4-1", FMT_PCT)
    ws["A27"].font = Font(name="Arial", bold=True)
    ws["I27"].font = Font(name="Arial", bold=True, color=BLACK)

    section(ws, 31, "Current-price reverse valuation", 6)
    header_row(ws, 32, ["Normalized P/E", "Implied EPS", "Implied net income", "Revenue @30% NM", "Revenue @35% NM", "Revenue @40% NM"])
    for r, multiple in enumerate([12.0, 14.0, 16.0], 33):
        put(ws, r, 1, multiple, FMT_MULT)
        put(ws, r, 2, f"=$B$4/A{r}", FMT_EPS)
        put(ws, r, 3, f"=B{r}*Assumptions!$B$25", FMT_MM)
        put(ws, r, 4, f"=C{r}/30%", FMT_MM)
        put(ws, r, 5, f"=C{r}/35%", FMT_MM)
        put(ws, r, 6, f"=C{r}/40%", FMT_MM)
    ws["A37"] = "Interpretation"
    ws["B37"] = "At 14x and 35%-40% normalized net margin, the current price implies sustainable revenue of roughly $41.5-$47.4bn, more than 2x FY2026E guidance-midpoint revenue."
    ws.merge_cells("B37:I37")
    ws["B37"].alignment = Alignment(wrap_text=True)
    ws["B37"].font = Font(name="Arial", italic=True, color=GRAY)
    ws.row_dimensions[37].height = 42
    finish_sheet(ws, {1: 34, 2: 16, 3: 17, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18, 9: 20}, "B4")


def build_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    add_title(ws, "Source manifest", "Evidence hierarchy: E0 filing/official observed data; E1 issuer primary; E3 reputable third-party estimate. Retrieval cutoff 2026-07-17 12:23 ET.", 8)
    headers = ["Source ID", "Publisher", "Title", "Published", "Tier", "Metric / use", "URL", "Limitations"]
    header_row(ws, 3, headers)

    rows = []
    if SOURCE_PATH.exists():
        with SOURCE_PATH.open(newline="", encoding="utf-8-sig") as handle:
            for item in csv.DictReader(handle):
                rows.append(
                    (
                        item.get("source_id") or item.get("Source ID") or "",
                        item.get("publisher") or item.get("Publisher") or "",
                        item.get("title") or item.get("Title") or "",
                        item.get("publication_date") or item.get("Published") or "",
                        item.get("evidence_tier") or item.get("tier") or item.get("Tier") or "",
                        item.get("metric") or item.get("Metric / use") or item.get("metric_or_use") or "",
                        item.get("stable_url") or item.get("url") or item.get("URL") or "",
                        item.get("limitations") or item.get("Limitations") or "",
                    )
                )
    valid_tiers = {"E0", "E1", "E2", "E3", "E4"}
    if len(rows) < 12 or any(not str(row[6]).startswith("http") or row[4] not in valid_tiers for row in rows):
        rows = INLINE_SOURCES

    for r, values in enumerate(rows, 4):
        for c, value in enumerate(values, 1):
            cell = put(ws, r, c, value)
            if c == 7 and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            elif c in (6, 8):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
    finish_sheet(ws, {1: 12, 2: 18, 3: 38, 4: 14, 5: 10, 6: 38, 7: 54, 8: 45}, "A4")
    ws.auto_filter.ref = f"A3:H{ws.max_row}"


def build_summary(wb: Workbook, positions: dict[str, dict[str, int]]) -> None:
    ws = wb.create_sheet("Summary", 0)
    add_title(ws, "SanDisk (NASDAQ: SNDK) financial model", "FY2026E-FY2030E | USD | cutoff 2026-07-17 12:23 ET intraday | status: screen-grade", 12)

    labels = [
        (4, "Research status", "Screen-grade"),
        (5, "Information cutoff", "2026-07-17 12:23 ET"),
        (6, "Current price", "=Assumptions!B4"),
        (7, "Market capitalization", "=Valuation!B6"),
        (8, "Enterprise value", "=Valuation!B9"),
    ]
    for row, label, value in labels:
        put(ws, row, 1, label)
        fmt = FMT_PRICE if row == 6 else (FMT_MM if row in (7, 8) else None)
        put(ws, row, 2, value, fmt)
    ws["A4"].font = Font(name="Arial", bold=True, color=NAVY)
    ws["B4"].fill = PatternFill("solid", fgColor=PALE_YELLOW)

    section(ws, 10, "Base-case operating forecast", 6)
    header_row(ws, 11, ["Metric", "FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"])
    summary_metrics = [("Revenue", 4, FMT_MM), ("Net income", 12, FMT_MM), ("Diluted EPS", 14, FMT_EPS), ("Gross margin", 6, FMT_PCT)]
    for r, (label, forecast_row, fmt) in enumerate(summary_metrics, 12):
        put(ws, r, 1, label)
        for col in range(2, 7):
            fcol = get_column_letter(col)
            put(ws, r, col, f"=Forecast!{fcol}{forecast_row}", fmt)

    section(ws, 18, "Scenario valuation", 6)
    header_row(ws, 19, ["Scenario", "Probability", "FY27 EPS", "FY29-30 avg EPS", "Fair value/share", "vs. price"])
    valuation_rows = {"Bear": 24, "Base": 25, "Bull": 26}
    for r, scenario in enumerate(["Bear", "Base", "Bull"], 20):
        vrow = valuation_rows[scenario]
        put(ws, r, 1, scenario)
        put(ws, r, 2, f"=Valuation!B{vrow}", FMT_PCT)
        put(ws, r, 3, f"=Valuation!C{vrow}", FMT_EPS)
        put(ws, r, 4, f"=Valuation!E{vrow}", FMT_EPS)
        put(ws, r, 5, f"=Valuation!H{vrow}", FMT_PRICE)
        put(ws, r, 6, f"=E{r}/$B$6-1", FMT_PCT)
    put(ws, 23, 1, "Probability-weighted")
    put(ws, 23, 5, "=Valuation!I27", FMT_PRICE)
    put(ws, 23, 6, "=Valuation!I28", FMT_PCT)
    ws["A23"].font = Font(name="Arial", bold=True)
    ws["E23"].font = Font(name="Arial", bold=True, color=GREEN_LINK)

    section(ws, 26, "Investor read-through", 6)
    notes = [
        "FY2026E is not reported actual: it combines FY2026 9M actuals with management's Q4 guidance midpoint.",
        "Q3 growth was primarily price/mix driven; the model explicitly normalizes ASP and gross margin after FY2027.",
        "The current price embeds either a low multiple on peak FY2027 earnings or a durable $41.5-$47.4bn normalized-revenue outcome.",
        "RPO and NBM contracts improve visibility but do not eliminate delivery, repricing, customer and supply risks.",
        "BiCS10 is stage-gated: no material base-case revenue before FY2028; HBF is excluded from base-case revenue.",
    ]
    for r, note in enumerate(notes, 27):
        put(ws, r, 1, f"{r - 26}.")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        put(ws, r, 2, note)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 30

    chart = LineChart()
    chart.title = "Scenario revenue path"
    chart.y_axis.title = "$mm"
    chart.x_axis.title = "Fiscal year"
    chart.height = 8
    chart.width = 14
    data = Reference(wb["Scenarios"], min_col=2, max_col=5, min_row=12, max_row=14)
    cats = Reference(wb["Scenarios"], min_col=2, max_col=5, min_row=11, max_row=11)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)
    for series, title in zip(chart.series, ["Bear", "Base", "Bull"]):
        series.graphicalProperties.line.width = 24000
        series.title = SeriesLabel(v=title)
    ws.add_chart(chart, "H10")

    finish_sheet(ws, {1: 25, 2: 18, 3: 18, 4: 18, 5: 18, 6: 18, 7: 3, 8: 14, 9: 14, 10: 14, 11: 14, 12: 14}, "B11")


def apply_tab_colors(wb: Workbook) -> None:
    colors = {
        "Summary": NAVY,
        "Historicals": "5B9BD5",
        "Drivers": "ED7D31",
        "Quarterly": "A5A5A5",
        "Forecast": "70AD47",
        "Scenarios": "FFC000",
        "Valuation": "7030A0",
        "Assumptions": "4472C4",
        "Sources": "264478",
    }
    for name, color in colors.items():
        wb[name].sheet_properties.tabColor = color


def load_and_validate_inputs() -> dict:
    """Load the frozen JSON pack and fail fast if code and evidence drift apart."""
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Missing frozen model input pack: {INPUT_PATH}")
    with INPUT_PATH.open(encoding="utf-8") as handle:
        inputs = json.load(handle)
    if inputs.get("contract_id") != "SNDK_2026-07-17T122300-0400":
        raise ValueError("model_inputs.json does not match the frozen research cutoff")
    if abs(inputs["market"]["price"] - 1470.63) > 1e-9 or inputs["market"]["price_type"] != "intraday":
        raise ValueError("Market observation drifted from the frozen intraday quote")
    if abs(inputs["fy26e_inputs"]["annual_estimate"]["revenue"] - 19283.0) > 1e-9:
        raise ValueError("FY2026E starting revenue drifted from the 9M-plus-guidance bridge")

    key_map = {
        "bit_growth": "bit_growth",
        "asp_index": "asp_index",
        "cost_index": "cost_index",
        "gross_margin": "gross_margin",
        "opex": "operating_expenses",
        "other_income": "other_income",
        "tax_rate": "tax_rate",
        "shares": "diluted_shares",
        "dc_share": "datacenter_share",
        "consumer_share": "consumer_share",
    }
    years = ["FY2027", "FY2028", "FY2029", "FY2030"]
    for scenario, expected in SCENARIOS.items():
        actual = inputs["scenarios"][scenario]
        if abs(actual["probability"] - expected["probability"]) > 1e-9:
            raise ValueError(f"Probability drift for {scenario}")
        for index, year in enumerate(years):
            annual = actual["annual"][year]
            for local_key, json_key in key_map.items():
                if abs(annual[json_key] - expected[local_key][index]) > 1e-9:
                    raise ValueError(f"Driver drift for {scenario} {year} {json_key}")
    return inputs


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    inputs = load_and_validate_inputs()
    wb = Workbook()
    wb.remove(wb.active)
    build_assumptions(wb)
    build_historicals(wb, inputs)
    driver_rows = build_drivers(wb)
    scenario_positions = build_scenarios(wb, driver_rows)
    build_forecast(wb, driver_rows)
    build_quarterly(wb, driver_rows, scenario_positions)
    build_valuation(wb, scenario_positions)
    build_sources(wb)
    build_summary(wb, scenario_positions)
    apply_tab_colors(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.properties.creator = "Codex"
    wb.properties.title = "SanDisk FY2026E-FY2030E financial model"
    wb.properties.subject = "AI hardware and NAND financial forecasting"
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
